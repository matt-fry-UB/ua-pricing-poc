"""
UA Simplified Pricing QA Checklist - populate, sync & migrate (v3)
==================================================================
Structure: one parent row per location (primary column holds the location
name) carrying SP Launch Date, QA By, QA Timestamp, QA Status, and a
% Complete formula; 23 child rows per location carrying the QA items in
the "QA Item" column (primary column blank on child rows), with Reference
Values from the "Simplified Pricing Rollout (2026)" sheet and Birthday Grid.

Modes
-----
  python sp_qa_checklist_sync.py populate                      # DRY RUN
  python sp_qa_checklist_sync.py populate --execute            # add missing locations
  python sp_qa_checklist_sync.py sync                          # DRY RUN
  python sp_qa_checklist_sync.py sync --execute                # update changed values
  python sp_qa_checklist_sync.py migrate-task-column           # DRY RUN
  python sp_qa_checklist_sync.py migrate-task-column --execute # one-time migration:
      moves QA item names from the primary (Task) column into the QA Item
      column on child rows, blanks the primary cell on child rows, and
      updates every parent's % Complete formula to count via QA Item.
  python sp_qa_checklist_sync.py stamp-locations               # DRY RUN
  python sp_qa_checklist_sync.py stamp-locations --execute     # backfill the
      "Location (ref)" column on every child row with its parent's location
      name (idempotent). populate stamps new locations automatically, so
      this is only for the one-time backfill of existing rows.
  python sp_qa_checklist_sync.py populate-slugs                # DRY RUN
  python sp_qa_checklist_sync.py populate-slugs --execute      # write WP Slug
      and CC Slug onto every row (parent + child) from a local Excel file
      (Location / WP Slug / CC Slug columns); child rows take their parent's
      location. Matches by location name, idempotent, reports any location
      that exists on one side only.
      Override the source file with --excel "C:\\path\\to\\file.xlsx".
      Needs openpyxl (pip install openpyxl).
  python sp_qa_checklist_sync.py apply-verified               # DRY RUN
  python sp_qa_checklist_sync.py apply-verified --execute     # write the
      Verified RYG formula to every parent row (Green = all items Done with
      no Mismatch flags; Red = all Done but a Mismatch exists; Yellow = in
      progress; blank = not started). populate sets it on new locations
      automatically; this backfills existing parents. Safe to re-run.
  python sp_qa_checklist_sync.py apply-template               # DRY RUN
  python sp_qa_checklist_sync.py apply-template --execute     # replicate the
      Akron template to every location: add missing QA Item rows (e.g.
      "Adventure 4 All") and backfill Edit Location, Notes / Instructions
      (GTM rows), and CC URL Slug (GTM rows, set to =[CC Slug]@row).
      Idempotent; never modifies Detail on existing rows and never touches
      Done, QA By, QA Timestamp, Command Center, Website, Mismatch, or
      Reference Value (the comparison tool owns those).
  python sp_qa_checklist_sync.py apply-hyperlinks             # DRY RUN
  python sp_qa_checklist_sync.py apply-hyperlinks --execute   # copy cell
      hyperlinks from the reference location (Akron) to the matching QA Item
      rows of every other location, for the Edit Location and Notes /
      Instructions columns. Smartsheet allows one hyperlink per cell.
      Idempotent; writes only the cell value + hyperlink.
  Add --cell-links to populate/sync to convert direct-source Reference
  Value cells (and parent SP Launch Date) into live Smartsheet cell links.
  Go-Karts Yes/No and the Birthday Promo lookup stay script-synced.

Auth: set SMARTSHEET_ACCESS_TOKEN in your environment.
"""

import argparse
import os
import sys
import time

import requests

# ============================================================
# CONFIG
# ============================================================
API_BASE = "https://api.smartsheet.com/2.0"
TOKEN_ENV_VAR = "SMARTSHEET_ACCESS_TOKEN"

QA_SHEET_ID = 3089334252031876        # UA Simplified Pricing QA Checklist
ROLLOUT_SHEET_ID = 998938260754308    # Simplified Pricing Rollout (2026)
BIRTHDAY_GRID_SHEET_ID = 2690238114647940  # 2026 Simplified Pricing Birthday Grid Discounts

# populate-slugs source: local Excel with Location / WP Slug / CC Slug columns.
# Override at runtime with --excel "C:\\path\\to\\file.xlsx".
SLUG_EXCEL_PATH = r"C:\Users\MattFry\Box\Matt Fry\Working\UA\UA-2026-memorial-day-urls.xlsx"
SLUG_EXCEL_COLS = {"location": "Location", "wp": "WP Slug", "cc": "CC Slug"}
# QA sheet target columns (resolved by title at runtime, not by ID).
SLUG_TARGET_COLS = {"wp": "WP Slug", "cc": "CC Slug"}

# QA sheet column IDs
QA_COL = {
    "task": 3847554830995332,         # primary column: location names only
    "launch_date": 4366082105446276,
    "qa_by": 8869681732816772,
    "qa_timestamp": 73588710608772,
    "qa_status": 4577188337979268,
    "pct": 2325388524294020,
    "done": 8351154458365828,
    "qa_item": 3615029596557188,      # QA checklist item names (child rows)
    "location_ref": 5187551106469764, # parent location name, repeated on each child row
    "detail": 1032805063888772,
    "ref": 5536404691259268,
    "source": 3284604877574020,
    "notes": 7788204504944516,
}

STATUS_FORMULA = ('=IF(COUNT(CHILDREN([QA Item]@row)) = 0, "", '
                  'IF(COUNTIF(CHILDREN(Done@row), 1) = 0, "\U0001F534 Not Started", '
                  'IF(COUNTIF(CHILDREN(Done@row), 1) >= COUNT(CHILDREN([QA Item]@row)), '
                  '"\U0001F7E2 Complete", "\U0001F7E1 In Progress")))')
PCT_FORMULA = ('=IF(COUNT(CHILDREN([QA Item]@row)) = 0, "", '
               'ROUND(COUNTIF(CHILDREN(Done@row), 1) / '
               'COUNT(CHILDREN([QA Item]@row)) * 100) + "%")')
# Verified RYG symbol: Green = all items Done and no Mismatch flags; Red = all
# Done but at least one Mismatch; Yellow = in progress; "" = not started.
VERIFIED_COL_TITLE = "Verified"
VERIFIED_FORMULA = ('=IF(COUNT(CHILDREN([QA Item]@row)) = 0, "", '
                    'IF(COUNTIF(CHILDREN(Done@row), 1) < COUNT(CHILDREN([QA Item]@row)), '
                    'IF(COUNTIF(CHILDREN(Done@row), 1) = 0, "", "Yellow"), '
                    'IF(COUNTIF(CHILDREN(Mismatch@row), <>"") = 0, "Green", "Red")))')

# --- Akron template additions: new columns (resolved by title) ---------------
EDIT_LOC_COL_TITLE = "Edit Location"
CC_URL_SLUG_COL_TITLE = "CC URL Slug"
NOTES_INSTR_COL_TITLE = "Notes / Instructions"

# Edit Location value shared by all promo items (and Adventure 4 All).
EDIT_LOC_PROMO = ("Location+Bday: https://www.urbanair.com/wp-admin/edit.php?post_type=promotion\n"
                  "Offers: https://www.urbanair.com/wp-admin/edit.php?post_type=promotions")
# Notes / Instructions text for the two GTM rows (constant across locations).
NOTES_TRIGGERS = ("Add CC URL Slug (from the column to the left) to page path regex in both "
                  "GTM triggers listed below.\nGTM Triggers Page:\n"
                  "https://tagmanager.google.com/#/container/accounts/2103147720/containers/"
                  "57046153/workspaces/231/triggers\nTrigger Names:\n"
                  "[UA Exp] - Bday Ultimate to Unlimited (Dom)\n"
                  "[UA Exp] - Bday Ultimate to Unlimited (Hist)")
NOTES_TAGS = ("Add CC URL Slug (from the column to the left) to fiveKidLocations list inside "
              "the GTM custom HTML tag listed below.\nGTM Tags Page:\n"
              "https://tagmanager.google.com/#/container/accounts/2103147720/containers/"
              "57046153/workspaces/231/tags\nTag Name:\n[UA Exp] - Bday Small Squad Party")

ADVENTURE_4_ALL_DETAIL = ('Adventure 4 All promo updated to simplified pricing '
                          '(legacy promo: gif image, "(4) Unlimited Play Tickets"; '
                          'simplified promo: jpg image, "(4) Unlimited Play+ Tickets")')

# Rollout sheet column IDs — immune to column renames.
# Run `list-rollout-cols` to verify or update these if columns are deleted/recreated.
ROLLOUT_COLS = {
    "location":            8959216101658500,   # Location Name
    "due":                   57569963184004,   # Due Date for Pricing Update
    "up_ticket":            747489107152772,   # Unlimited Play Ticket
    "go_karts":            4371379188436868,   # Go-Karts
    "up_plus":             6349633424232324,   # Unlimited Play + Ticket
    "parent_pass":         8249331818991492,   # Parent Pass
    "shorty40":            3745732191620996,   # Shorty 40
    "tier":                4839224993156996,   # BIRTHDAY TIER
    "up_party":            8601433237917572,   # Unlimited Play Party
    "ssp_5x":              3253408680415108,   # NEW Small Squad Party 2.0 5X Guests
    "room":                1493932377935748,   # Room Upgrade
    "suite":               5997532005306244,   # Suite Upgrade
    "up_membership":        930982424514436,   # Unlimited Play 12-Month MTM Membership
    "parent_membership":   5434582051884932,   # Parent Membership
    "shorty40_membership": 3182782238199684,   # Shorty 40 Membership
}

GRID_COLS = {"tier": "Birthday Tier", "promo": "Promo Name"}

# Checklist template (24 items): (item, detail, rollout_key or None, source label or None)
CHECKLIST = [
    ("Unlimited Play Ticket", "Price in website and CC match", "up_ticket", "Unlimited Play Ticket"),
    ("Has Unlimited Play +?", "Does this location have Unlimited Play +?", "go_karts", "Go-Karts"),
    ("Unlimited Play + Ticket", "If location has Unlimited Play +, make sure price in website and CC match", "up_plus", "Unlimited Play + Ticket"),
    ("Parent Pass", "Price in website and CC match", "parent_pass", "Parent Pass"),
    ("Shorty 40", "Price in website and CC match", "shorty40", "Shorty 40"),
    ("Socks", "Price in website and CC match", None, None),
    ("Ticket Attractions", "List in website and CC match", None, None),
    ("Unlimited Play Birthday", "Price in website and CC match", "up_party", "Unlimited Play Party"),
    ("Birthday Attractions", "List in website and CC match", None, None),
    ("Private Room", "Price in website and CC match", "room", "Room Upgrade"),
    ("Private Room Amenities", "List in website and CC match", None, None),
    ("VIP Suite", "Price in website and CC match", "suite", "Suite Upgrade"),
    ("VIP Suite Amenities", "List in website and CC match", None, None),
    ("Unlimited Play Membership", "Price in website and CC match", "up_membership", "Unlimited Play 12-Month MTM Membership"),
    ("Shorty 40 Membership", "Price in website and CC match", "shorty40_membership", "Shorty 40 Membership"),
    ("Parent Pass Membership", "Price in website and CC match", "parent_membership", "Parent Membership"),
    ("Membership Attractions", "List in website and CC match", None, None),
    ("Ultimate to Unlimited (GTM)", "CC says Unlimited and Small Squad", None, None),
    ("Small Squad Pop Up (GTM)", "CC SSP popup says 5 kids", None, None),
    ("$100/25% Off Birthday Promo", "Moved to new simplified pricing bday discount promo", None, None),
    ("$100/25% Off Birthday Promo Discount", "Matches Smartsheet birthday grid tier and price listed in CC", "PROMO", "Birthday Grid Promo Name"),
    ("Small Squad Promo", "Moved to new simplified pricing SSP promo (5 kids, no pizza, shared party host)", None, None),
    ("Small Squad Promo Price", "Price matches column in simplified pricing rollout smartsheet and price listed in CC", "ssp_5x", "Small Squad Party 5X Guests"),
    ("Adventure 4 All", ADVENTURE_4_ALL_DETAIL, None, None),
]

# Per-item template values for the Akron-added columns, keyed by QA Item.
# (edit_location, notes_instructions or None, needs_cc_url_slug)
GROUP_TICKET = ("Unlimited Play Ticket", "Has Unlimited Play +?", "Unlimited Play + Ticket",
                "Parent Pass", "Shorty 40", "Socks", "Ticket Attractions")
GROUP_BIRTHDAY = ("Unlimited Play Birthday", "Birthday Attractions", "Private Room",
                  "Private Room Amenities", "VIP Suite", "VIP Suite Amenities")
GROUP_MEMBERSHIP = ("Unlimited Play Membership", "Shorty 40 Membership",
                    "Parent Pass Membership", "Membership Attractions")
GROUP_PROMO = ("$100/25% Off Birthday Promo", "$100/25% Off Birthday Promo Discount",
               "Small Squad Promo", "Small Squad Promo Price", "Adventure 4 All")
TEMPLATE_FIELDS = {}
for _i in GROUP_TICKET:
    TEMPLATE_FIELDS[_i] = ("Ticket", None, False)
for _i in GROUP_BIRTHDAY:
    TEMPLATE_FIELDS[_i] = ("Birthday", None, False)
for _i in GROUP_MEMBERSHIP:
    TEMPLATE_FIELDS[_i] = ("Membership", None, False)
for _i in GROUP_PROMO:
    TEMPLATE_FIELDS[_i] = (EDIT_LOC_PROMO, None, False)
TEMPLATE_FIELDS["Ultimate to Unlimited (GTM)"] = ("GTM Triggers", NOTES_TRIGGERS, True)
TEMPLATE_FIELDS["Small Squad Pop Up (GTM)"] = ("GTM Tags", NOTES_TAGS, True)

CELL_LINKABLE = {"up_ticket", "up_plus", "parent_pass", "shorty40", "up_party",
                 "ssp_5x", "room", "suite", "up_membership", "parent_membership",
                 "shorty40_membership"}

REQUEST_PAUSE_SEC = 0.4
UPDATE_BATCH = 400
MAX_RETRIES = 5

# ============================================================
# API helpers
# ============================================================

def _headers():
    token = os.environ.get(TOKEN_ENV_VAR)
    if not token:
        sys.exit(f"ERROR: set {TOKEN_ENV_VAR} in your environment.")
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


def api(method, path, **kwargs):
    url = f"{API_BASE}{path}"
    for attempt in range(MAX_RETRIES):
        resp = requests.request(method, url, headers=_headers(), timeout=120, **kwargs)
        if resp.status_code == 429 or resp.status_code >= 500:
            wait = 2 ** attempt
            print(f"  ... {resp.status_code} from API, retrying in {wait}s")
            time.sleep(wait)
            continue
        if not resp.ok:
            sys.exit(f"ERROR {resp.status_code} on {method} {path}: {resp.text[:500]}")
        return resp.json()
    sys.exit(f"ERROR: gave up after {MAX_RETRIES} retries on {method} {path}")


def get_sheet(sheet_id):
    return api("GET", f"/sheets/{sheet_id}?pageSize=10000")


def put_rows_batched(updates):
    for i in range(0, len(updates), UPDATE_BATCH):
        api("PUT", f"/sheets/{QA_SHEET_ID}/rows", json=updates[i:i + UPDATE_BATCH])
        time.sleep(REQUEST_PAUSE_SEC)


# ============================================================
# Data shaping
# ============================================================

def col_map(sheet):
    return {c["title"]: c["id"] for c in sheet["columns"]}


def row_values(sheet):
    out = []
    for r in sheet["rows"]:
        d = {"_row_id": r["id"], "_parent_id": r.get("parentId")}
        for c in r.get("cells", []):
            d[c["columnId"]] = c.get("value")
        out.append(d)
    return out


def load_birthday_grid():
    sheet = get_sheet(BIRTHDAY_GRID_SHEET_ID)
    cols = col_map(sheet)
    tier_col, promo_col = cols[GRID_COLS["tier"]], cols[GRID_COLS["promo"]]
    grid = {}
    for r in row_values(sheet):
        tier, promo = r.get(tier_col), r.get(promo_col)
        if tier is None or promo is None:
            continue
        tier = int(float(tier))
        if isinstance(promo, (int, float)):
            promo = f"{promo:.0%} Off"  # 0.25 -> "25% Off"
        grid[tier] = str(promo)
    return grid


def parse_tier(tier_label):
    if not tier_label:
        return None
    try:
        return int(str(tier_label).split("-")[0].strip())
    except ValueError:
        return None


def load_rollout_locations():
    sheet = get_sheet(ROLLOUT_SHEET_ID)
    cols = col_map(sheet)
    key_to_colid = {}
    for k, v in ROLLOUT_COLS.items():
        if isinstance(v, int):
            key_to_colid[k] = v  # already a column ID — rename-resistant
        elif v in cols:
            key_to_colid[k] = cols[v]
        else:
            sys.exit(
                f"ERROR: rollout sheet column '{v}' (key '{k}') not found.\n"
                f"Available columns: {sorted(cols)}\n"
                f"Update ROLLOUT_COLS['{k}'] to the correct title, or replace it with "
                f"the integer column ID (run 'list-rollout-cols' to see all IDs)."
            )
    locations, order = {}, []
    for r in row_values(sheet):
        name = r.get(key_to_colid["location"])
        if not name:
            continue
        data = {k: r.get(cid) for k, cid in key_to_colid.items()}
        data["_src_row_id"] = r["_row_id"]
        data["_src_col_ids"] = key_to_colid
        if name in locations:
            prev = locations[name]
            if prev.get("due") != data.get("due"):
                prev["_due_conflict"] = sorted(
                    str(d) for d in {prev.get("due"), data.get("due")} if d)
            continue
        locations[name] = data
        order.append(name)
    return [(n, locations[n]) for n in order]


def ref_value_for(key, data, grid):
    if key == "go_karts":
        return ("Yes" if data.get("go_karts") else "No"), None
    if key == "PROMO":
        tier_n = parse_tier(data.get("tier"))
        if tier_n is None:
            return None, "No BIRTHDAY TIER set in rollout sheet"
        promo = grid.get(tier_n)
        if promo is None:
            return None, f"Tier {tier_n} not present in Birthday Grid"
        return promo, None
    val = data.get(key)
    return (val, None) if val is not None else (None, None)


def source_label_for(key, data, base_label):
    if key == "PROMO":
        return f"{base_label} (via BIRTHDAY TIER {data.get('tier')})"
    return base_label


def child_item_name(kid):
    """QA item name for a child row: QA Item column, falling back to the
    primary column for any rows that predate the column migration."""
    return kid.get(QA_COL["qa_item"]) or kid.get(QA_COL["task"])


def build_parent_cells(name, data, verified_col_id=None):
    cells = [
        {"columnId": QA_COL["task"], "value": name},
        {"columnId": QA_COL["qa_status"], "formula": STATUS_FORMULA},
        {"columnId": QA_COL["pct"], "formula": PCT_FORMULA},
    ]
    if verified_col_id:
        cells.append({"columnId": verified_col_id, "formula": VERIFIED_FORMULA})
    notes = []
    if data.get("_due_conflict"):
        notes.append("Rollout sheet has duplicate rows for this location with "
                     "different due dates (" + " / ".join(data["_due_conflict"]) +
                     "); confirm and dedupe.")
    elif data.get("due"):
        cells.append({"columnId": QA_COL["launch_date"], "value": str(data["due"])})
    else:
        notes.append("No due date set in rollout sheet")
    if notes:
        cells.append({"columnId": QA_COL["notes"], "value": " | ".join(notes)})
    return cells


def build_child_rows(parent_id, name, data, grid):
    rows = []
    for item, detail, key, src in CHECKLIST:
        cells = [
            {"columnId": QA_COL["qa_item"], "value": item},
            {"columnId": QA_COL["location_ref"], "value": name},
            {"columnId": QA_COL["detail"], "value": detail},
        ]
        if key:
            cells.append({"columnId": QA_COL["source"],
                          "value": source_label_for(key, data, src)})
            val, note = ref_value_for(key, data, grid)
            if val is not None:
                cells.append({"columnId": QA_COL["ref"], "value": val})
            if note:
                cells.append({"columnId": QA_COL["notes"], "value": note})
        rows.append({"parentId": parent_id, "toBottom": True, "cells": cells})
    return rows


# ============================================================
# Modes
# ============================================================

def existing_sections(qa_sheet):
    parents, children = {}, []
    task_col = QA_COL["task"]
    for r in row_values(qa_sheet):
        if r["_parent_id"] is None:
            name = r.get(task_col)
            if name:
                parents[str(name)] = r["_row_id"]
        else:
            children.append(r)
    return parents, children


def mode_populate(execute, use_links):
    grid = load_birthday_grid()
    rollout = load_rollout_locations()
    qa_sheet = get_sheet(QA_SHEET_ID)
    parents, _ = existing_sections(qa_sheet)
    verified_col_id = col_map(qa_sheet).get(VERIFIED_COL_TITLE)

    missing = [(n, d) for n, d in rollout if n not in parents]
    print(f"Rollout locations (deduped): {len(rollout)}")
    print(f"Already in QA sheet:         {len(parents)}")
    print(f"To add:                      {len(missing)}")
    for n, _d in missing[:15]:
        print(f"  + {n}")
    if len(missing) > 15:
        print(f"  ... and {len(missing) - 15} more")
    if not execute:
        print("\nDRY RUN. Re-run with --execute to write.")
        return

    data_by_parent = {}
    for i, (name, data) in enumerate(missing, 1):
        res = api("POST", f"/sheets/{QA_SHEET_ID}/rows",
                  json=[{"toBottom": True, "cells": build_parent_cells(name, data, verified_col_id)}])
        result = res["result"]
        parent_id = result[0]["id"] if isinstance(result, list) else result["id"]
        time.sleep(REQUEST_PAUSE_SEC)
        api("POST", f"/sheets/{QA_SHEET_ID}/rows",
            json=build_child_rows(parent_id, name, data, grid))
        data_by_parent[parent_id] = data
        print(f"  [{i}/{len(missing)}] added {name}")
        time.sleep(REQUEST_PAUSE_SEC)

    if use_links:
        _apply_links(data_by_parent)
    print("Done.")


def mode_sync(execute, use_links):
    grid = load_birthday_grid()
    rollout = dict(load_rollout_locations())
    qa_sheet = get_sheet(QA_SHEET_ID)
    parents, children = existing_sections(qa_sheet)

    parent_to_name = {rid: name for name, rid in parents.items()}
    kids_by_parent = {}
    for k in children:
        kids_by_parent.setdefault(k["_parent_id"], []).append(k)

    all_rows = {r["_row_id"]: r for r in row_values(qa_sheet)}
    task_to_key = {t: k for t, _, k, _ in CHECKLIST if k}
    updates, diffs = [], []

    # Parent-level: SP Launch Date
    for name, parent_id in parents.items():
        data = rollout.get(name)
        if data is None:
            diffs.append(f"  ! '{name}' not found in rollout sheet (skipped)")
            continue
        if data.get("_due_conflict"):
            continue
        old = all_rows.get(parent_id, {}).get(QA_COL["launch_date"])
        new = str(data["due"]) if data.get("due") else None
        if _norm(old) != _norm(new):
            diffs.append(f"  {name} / SP Launch Date: {old!r} -> {new!r}")
            updates.append({"id": parent_id, "cells": [
                {"columnId": QA_COL["launch_date"],
                 "value": new if new is not None else ""}]})

    # Child-level Reference Values
    for parent_id, kids in kids_by_parent.items():
        name = parent_to_name.get(parent_id)
        data = rollout.get(name)
        if data is None:
            continue
        for kid in kids:
            key = task_to_key.get(str(child_item_name(kid)))
            if not key:
                continue
            new_val, _note = ref_value_for(key, data, grid)
            old_val = kid.get(QA_COL["ref"])
            if _norm(old_val) != _norm(new_val):
                diffs.append(f"  {name} / {child_item_name(kid)}: "
                             f"{old_val!r} -> {new_val!r}")
                updates.append({"id": kid["_row_id"], "cells": [
                    {"columnId": QA_COL["ref"],
                     "value": new_val if new_val is not None else ""}]})

    print(f"Changes: {len(updates)}")
    for line in diffs[:40]:
        print(line)
    if len(diffs) > 40:
        print(f"  ... and {len(diffs) - 40} more")
    if not execute:
        print("\nDRY RUN. Re-run with --execute to write.")
        return
    put_rows_batched(updates)
    if use_links:
        data_by_parent = {pid: rollout[parent_to_name[pid]]
                          for pid in kids_by_parent
                          if parent_to_name.get(pid) in rollout}
        _apply_links(data_by_parent, kids_by_parent)
    print("Done.")


def mode_migrate(execute):
    """One-time: move QA item names from the primary column to the QA Item
    column on child rows, blank the primary cell, refresh parent formulas."""
    qa_sheet = get_sheet(QA_SHEET_ID)
    rows = row_values(qa_sheet)
    task_col, item_col = QA_COL["task"], QA_COL["qa_item"]

    child_updates, parent_updates = [], []
    for r in rows:
        if r["_parent_id"] is None:
            parent_updates.append({"id": r["_row_id"], "cells": [
                {"columnId": QA_COL["pct"], "formula": PCT_FORMULA},
                {"columnId": QA_COL["qa_status"], "formula": STATUS_FORMULA}]})
        else:
            task_val, item_val = r.get(task_col), r.get(item_col)
            if task_val and not item_val:
                child_updates.append({"id": r["_row_id"], "cells": [
                    {"columnId": item_col, "value": task_val},
                    {"columnId": task_col, "value": ""}]})

    print(f"Child rows to migrate (Task -> QA Item, then blank Task): {len(child_updates)}")
    print(f"Parent rows to receive updated % Complete formula:        {len(parent_updates)}")
    if not execute:
        print("\nDRY RUN. Re-run with --execute to write.")
        return
    # Children first, then parent formulas, so the new formula counts
    # populated QA Item cells the moment it lands.
    put_rows_batched(child_updates)
    put_rows_batched(parent_updates)
    print("Done.")


def mode_stamp_locations(execute):
    """Backfill the 'Location (ref)' column on every child row with its
    parent's location name. Idempotent: skips children already stamped
    with the correct value."""
    qa_sheet = get_sheet(QA_SHEET_ID)
    parents, children = existing_sections(qa_sheet)
    parent_to_name = {rid: name for name, rid in parents.items()}
    loc_col = QA_COL["location_ref"]

    updates, skipped, orphan = [], 0, 0
    for kid in children:
        name = parent_to_name.get(kid["_parent_id"])
        if name is None:
            orphan += 1
            continue
        if kid.get(loc_col) == name:
            skipped += 1
            continue
        updates.append({"id": kid["_row_id"],
                        "cells": [{"columnId": loc_col, "value": name}]})

    print(f"Child rows needing a location stamp: {len(updates)}")
    print(f"Already correct (skipped):           {skipped}")
    if orphan:
        print(f"Children with no matching parent:    {orphan}")
    if not execute:
        print("\nDRY RUN. Re-run with --execute to write.")
        return
    put_rows_batched(updates)
    print("Done.")


def mode_populate_slugs(execute, excel_path):
    """Populate the WP Slug and CC Slug columns on every row (parent and
    child) from a local Excel file keyed by Location; child rows take their
    parent's location. Idempotent: skips rows already holding both correct
    slugs. Columns are resolved by title."""
    try:
        import openpyxl
    except ImportError:
        sys.exit("ERROR: this mode needs openpyxl. Run: pip install openpyxl")

    path = excel_path or SLUG_EXCEL_PATH
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    hdr = [c.value for c in ws[1]]
    try:
        li = hdr.index(SLUG_EXCEL_COLS["location"])
        wi = hdr.index(SLUG_EXCEL_COLS["wp"])
        ci = hdr.index(SLUG_EXCEL_COLS["cc"])
    except ValueError:
        sys.exit(f"ERROR: expected columns {list(SLUG_EXCEL_COLS.values())} in {path}; "
                 f"found {hdr}")

    slug_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        loc = (row[li] or "").strip() if row[li] else ""
        if not loc:
            continue
        slug_map[loc] = {"wp": (row[wi] or "").strip() if row[wi] else "",
                         "cc": (row[ci] or "").strip() if row[ci] else ""}

    qa_sheet = get_sheet(QA_SHEET_ID)
    cols = col_map(qa_sheet)
    try:
        wp_col, cc_col = cols[SLUG_TARGET_COLS["wp"]], cols[SLUG_TARGET_COLS["cc"]]
    except KeyError:
        sys.exit(f"ERROR: QA sheet is missing target columns "
                 f"{list(SLUG_TARGET_COLS.values())}; found {sorted(cols)}")
    parents, children = existing_sections(qa_sheet)
    parent_to_name = {rid: name for name, rid in parents.items()}
    all_rows = {r["_row_id"]: r for r in row_values(qa_sheet)}

    # Build (row_id, location_name) for every row: parents by their own name,
    # children by their parent's name.
    row_loc = [(rid, name) for name, rid in parents.items()]
    orphan = 0
    for kid in children:
        name = parent_to_name.get(kid["_parent_id"])
        if name is None:
            orphan += 1
            continue
        row_loc.append((kid["_row_id"], name))

    updates, skipped = [], 0
    sheet_unmatched_rows, used = {}, set()
    for rid, name in row_loc:
        slugs = slug_map.get(name)
        if slugs is None:
            sheet_unmatched_rows[name] = sheet_unmatched_rows.get(name, 0) + 1
            continue
        used.add(name)
        cur = all_rows.get(rid, {})
        cells = []
        if slugs["wp"] and cur.get(wp_col) != slugs["wp"]:
            cells.append({"columnId": wp_col, "value": slugs["wp"]})
        if slugs["cc"] and cur.get(cc_col) != slugs["cc"]:
            cells.append({"columnId": cc_col, "value": slugs["cc"]})
        if cells:
            updates.append({"id": rid, "cells": cells})
        else:
            skipped += 1

    excel_unmatched = [n for n in slug_map if n not in used]

    print(f"Excel locations:                 {len(slug_map)}")
    print(f"QA sheet locations (parents):    {len(parents)}")
    print(f"Total rows considered:           {len(row_loc)}")
    print(f"Rows to update:                  {len(updates)}")
    print(f"Already correct (skipped):       {skipped}")
    if orphan:
        print(f"Child rows with no parent:       {orphan}")
    if sheet_unmatched_rows:
        total = sum(sheet_unmatched_rows.values())
        print(f"\nSheet locations with NO Excel slug "
              f"({len(sheet_unmatched_rows)} locations, {total} rows):")
        for n in sheet_unmatched_rows:
            print(f"  - {n}")
    if excel_unmatched:
        print(f"\nExcel locations not found in sheet ({len(excel_unmatched)}):")
        for n in excel_unmatched:
            print(f"  - {n}")
    if not execute:
        print("\nDRY RUN. Re-run with --execute to write.")
        return
    put_rows_batched(updates)
    print("Done.")


def mode_apply_verified(execute):
    """Write the Verified RYG formula to every parent (location) row. Safe to
    re-run; it simply (re)sets the formula. The column is resolved by title."""
    qa_sheet = get_sheet(QA_SHEET_ID)
    verified_col_id = col_map(qa_sheet).get(VERIFIED_COL_TITLE)
    if not verified_col_id:
        sys.exit(f"ERROR: QA sheet has no '{VERIFIED_COL_TITLE}' column; "
                 f"add the RYG symbol column first.")
    parents, _ = existing_sections(qa_sheet)
    updates = [{"id": rid, "cells": [
        {"columnId": verified_col_id, "formula": VERIFIED_FORMULA}]}
        for rid in parents.values()]

    print(f"Parent rows to receive the Verified formula: {len(updates)}")
    if not execute:
        print("\nDRY RUN. Re-run with --execute to write.")
        return
    put_rows_batched(updates)
    print("Done.")


def mode_apply_template(execute):
    """Replicate the Akron template to every location: add any missing QA Item
    rows (e.g. 'Adventure 4 All') and backfill the Edit Location, Notes /
    Instructions (GTM rows), and CC URL Slug (GTM rows, set to the formula
    =[CC Slug]@row) columns. Idempotent. Never modifies Detail on existing
    rows, and never touches Done, QA By, QA Timestamp, Command Center,
    Website, Mismatch, or Reference Value (the comparison tool owns those)."""
    qa_sheet = get_sheet(QA_SHEET_ID)
    cols = col_map(qa_sheet)
    try:
        edit_col = cols[EDIT_LOC_COL_TITLE]
        ccurl_col = cols[CC_URL_SLUG_COL_TITLE]
        notes_col = cols[NOTES_INSTR_COL_TITLE]
    except KeyError as e:
        sys.exit(f"ERROR: QA sheet missing expected column {e}; found {sorted(cols)}")
    ccslug_col = cols.get("CC Slug")

    qa_item_col = QA_COL["qa_item"]
    detail_col = QA_COL["detail"]
    source_col = QA_COL["source"]
    locref_col = QA_COL["location_ref"]

    parents, children = existing_sections(qa_sheet)
    parent_to_name = {rid: name for name, rid in parents.items()}

    # parent_id -> {qa_item_name: child_row_dict}
    kids_by_parent = {}
    for k in children:
        name = parent_to_name.get(k["_parent_id"])
        if name is None:
            continue
        kids_by_parent.setdefault(k["_parent_id"], {})[str(k.get(qa_item_col))] = k

    # Source each item's Detail from an existing row in the sheet (e.g. Akron's
    # Adventure 4 All row), so added rows replicate what's actually there rather
    # than any hardcoded text. Falls back to the CHECKLIST default if absent.
    existing_detail_by_item = {}
    for k in children:
        item_name = str(k.get(qa_item_col))
        if item_name not in existing_detail_by_item and k.get(detail_col):
            existing_detail_by_item[item_name] = k.get(detail_col)

    CC_SLUG_FORMULA = "=[CC Slug]@row"

    updates, adds = [], []
    add_counts = {}
    for parent_id, name in parent_to_name.items():
        kids = kids_by_parent.get(parent_id, {})
        for item, detail, _key, src in CHECKLIST:
            edit_loc, notes, needs_cc = TEMPLATE_FIELDS.get(item, (None, None, False))
            if item in kids:
                kid = kids[item]
                cells = []
                # NOTE: Detail is intentionally NOT updated on existing rows.
                if edit_loc and kid.get(edit_col) != edit_loc:
                    cells.append({"columnId": edit_col, "value": edit_loc})
                if notes and kid.get(notes_col) != notes:
                    cells.append({"columnId": notes_col, "value": notes})
                # CC URL Slug formula is idempotent: once set it evaluates to the
                # row's CC Slug, so compare the displayed value to that to decide.
                if needs_cc and ccslug_col is not None:
                    if kid.get(ccurl_col) != kid.get(ccslug_col):
                        cells.append({"columnId": ccurl_col, "formula": CC_SLUG_FORMULA})
                if cells:
                    updates.append({"id": kid["_row_id"], "cells": cells})
            else:
                # Missing item (chiefly 'Adventure 4 All'): create the row.
                cells = [
                    {"columnId": qa_item_col, "value": item},
                    {"columnId": locref_col, "value": name},
                ]
                detail_to_use = existing_detail_by_item.get(item, detail)
                if detail_to_use:
                    cells.append({"columnId": detail_col, "value": detail_to_use})
                if src:
                    cells.append({"columnId": source_col, "value": src})
                if edit_loc:
                    cells.append({"columnId": edit_col, "value": edit_loc})
                if notes:
                    cells.append({"columnId": notes_col, "value": notes})
                if needs_cc:
                    cells.append({"columnId": ccurl_col, "formula": CC_SLUG_FORMULA})
                adds.append({"parentId": parent_id, "toBottom": True, "cells": cells})
                add_counts[item] = add_counts.get(item, 0) + 1

    print(f"Locations:                     {len(parent_to_name)}")
    print(f"Rows to ADD (missing items):   {len(adds)}")
    for item, n in add_counts.items():
        print(f"    + {item}: {n}")
    print(f"Existing rows to UPDATE:       {len(updates)}")
    if not execute:
        print("\nDRY RUN. Re-run with --execute to write.")
        return
    # Adds first (so the new rows exist), then the column backfills.
    # A single add-request can only target ONE parent (Smartsheet requires every
    # row in the request to share the same location), so POST one group per parent.
    adds_by_parent = {}
    for a in adds:
        adds_by_parent.setdefault(a["parentId"], []).append(a)
    for n, (pid, group) in enumerate(adds_by_parent.items(), 1):
        api("POST", f"/sheets/{QA_SHEET_ID}/rows", json=group)
        if n % 25 == 0:
            print(f"  ...added rows for {n}/{len(adds_by_parent)} locations")
        time.sleep(REQUEST_PAUSE_SEC)
    print(f"  added rows for {len(adds_by_parent)} locations; applying {len(updates)} updates")
    put_rows_batched(updates)
    print("Done.")


# Location used as the source of truth for the template (its cells are copied
# to every other location). Akron is the fully-built reference.
REFERENCE_LOCATION = "Akron, OH"
# Columns whose cell hyperlinks get replicated from the reference location.
HYPERLINK_COL_TITLES = (EDIT_LOC_COL_TITLE, NOTES_INSTR_COL_TITLE)


def mode_apply_hyperlinks(execute):
    """Copy cell hyperlinks from the reference location (Akron) to the matching
    QA Item rows of every other location, for the Edit Location and Notes /
    Instructions columns. Smartsheet allows one hyperlink per cell; whatever
    single hyperlink each reference cell holds is replicated verbatim.
    Idempotent: a target cell is skipped if it already links to the same URL.
    Only the cell value + hyperlink are written; nothing else is touched."""
    qa_sheet = get_sheet(QA_SHEET_ID)
    cols = col_map(qa_sheet)
    link_cols = {cols[t]: t for t in HYPERLINK_COL_TITLES if t in cols}
    if not link_cols:
        sys.exit("ERROR: none of the hyperlink columns were found on the sheet.")
    qa_item_col = QA_COL["task"]   # primary col holds location name on parents
    item_col = QA_COL["qa_item"]

    # Parse rows keeping full cell objects (value + hyperlink).
    parents, rows_by_id = {}, {}
    for r in qa_sheet["rows"]:
        cellmap = {c["columnId"]: c for c in r.get("cells", [])}
        rows_by_id[r["id"]] = {"parent_id": r.get("parentId"), "cells": cellmap}
        if r.get("parentId") is None:
            nm = cellmap.get(qa_item_col, {}).get("value")
            if nm:
                parents[str(nm)] = r["id"]
    parent_to_name = {rid: n for n, rid in parents.items()}

    # parent_id -> {qa_item_name: (row_id, cellmap)}
    kids_by_parent = {}
    for rid, info in rows_by_id.items():
        if info["parent_id"] is None:
            continue
        name = parent_to_name.get(info["parent_id"])
        if not name:
            continue
        item = str(info["cells"].get(item_col, {}).get("value"))
        kids_by_parent.setdefault(info["parent_id"], {})[item] = (rid, info["cells"])

    ref_pid = parents.get(REFERENCE_LOCATION)
    if not ref_pid:
        sys.exit(f"ERROR: reference location '{REFERENCE_LOCATION}' not found.")
    ref_kids = kids_by_parent.get(ref_pid, {})

    # Capture hyperlinks from the reference: {item: {col_id: (value, hyperlink)}}
    template_links = {}
    for item, (_rid, cells) in ref_kids.items():
        for col_id in link_cols:
            cell = cells.get(col_id)
            hl = cell.get("hyperlink") if cell else None
            if hl and hl.get("url"):
                template_links.setdefault(item, {})[col_id] = (cell.get("value"), hl)

    found = {link_cols[c]: 0 for c in link_cols}
    for d in template_links.values():
        for c in d:
            found[link_cols[c]] += 1
    print(f"Hyperlinks found on reference ({REFERENCE_LOCATION}):")
    for title, n in found.items():
        print(f"    {title}: {n} item row(s)")
    if not template_links:
        print("\nNo cell hyperlinks on the reference location's Edit Location /"
              " Notes / Instructions cells. If those URLs are clickable in the UI,"
              " they are auto-rendered plain text (already copied), not hyperlink"
              " objects, so there is nothing to replicate.")
        return

    updates = []
    for pid, name in parent_to_name.items():
        if pid == ref_pid:
            continue
        kids = kids_by_parent.get(pid, {})
        for item, coldata in template_links.items():
            target = kids.get(item)
            if not target:
                continue
            tgt_rid, tgt_cells = target
            payload = []
            for col_id, (val, hl) in coldata.items():
                cur = tgt_cells.get(col_id, {})
                cur_url = (cur.get("hyperlink") or {}).get("url")
                if cur_url == hl.get("url"):
                    continue  # idempotent
                new_hl = {"url": hl["url"]}
                if hl.get("label"):
                    new_hl["label"] = hl["label"]
                cell = {"columnId": col_id, "hyperlink": new_hl}
                if val is not None:
                    cell["value"] = val
                payload.append(cell)
            if payload:
                updates.append({"id": tgt_rid, "cells": payload})

    print(f"Target cells to update across other locations: {len(updates)} rows")
    if not execute:
        print("\nDRY RUN. Re-run with --execute to write.")
        return
    put_rows_batched(updates)
    print("Done.")


def make_cell_link_updates(kids_by_parent, data_by_parent):
    updates = []
    ref_col = QA_COL["ref"]
    task_to_key = {t: k for t, _, k, _ in CHECKLIST if k}
    for parent_id, data in data_by_parent.items():
        if data.get("due") and not data.get("_due_conflict"):
            updates.append({"id": parent_id, "cells": [{
                "columnId": QA_COL["launch_date"], "value": None,
                "linkInFromCell": {"sheetId": ROLLOUT_SHEET_ID,
                                   "rowId": data["_src_row_id"],
                                   "columnId": data["_src_col_ids"]["due"]}}]})
        for kid in kids_by_parent.get(parent_id, []):
            key = task_to_key.get(str(child_item_name(kid)))
            if key not in CELL_LINKABLE or data.get(key) is None:
                continue
            updates.append({"id": kid["_row_id"], "cells": [{
                "columnId": ref_col, "value": None,
                "linkInFromCell": {"sheetId": ROLLOUT_SHEET_ID,
                                   "rowId": data["_src_row_id"],
                                   "columnId": data["_src_col_ids"][key]}}]})
    return updates


def _apply_links(data_by_parent, kids_by_parent=None):
    if kids_by_parent is None:
        qa_sheet = get_sheet(QA_SHEET_ID)
        _, children = existing_sections(qa_sheet)
        kids_by_parent = {}
        for k in children:
            kids_by_parent.setdefault(k["_parent_id"], []).append(k)
    updates = make_cell_link_updates(kids_by_parent, data_by_parent)
    print(f"Creating {len(updates)} cell links...")
    for i in range(0, len(updates), 100):
        api("PUT", f"/sheets/{QA_SHEET_ID}/rows", json=updates[i:i + 100])
        time.sleep(REQUEST_PAUSE_SEC)


def _norm(v):
    if v is None or v == "":
        return None
    try:
        return round(float(v), 6)
    except (TypeError, ValueError):
        return str(v)


def mode_list_rollout_cols():
    sheet = get_sheet(ROLLOUT_SHEET_ID)
    print(f"Rollout sheet columns (copy an ID into ROLLOUT_COLS to make that key rename-resistant):\n")
    for c in sheet["columns"]:
        marker = ""
        for k, v in ROLLOUT_COLS.items():
            if v == c["title"] or v == c["id"]:
                marker = f"  <- ROLLOUT_COLS['{k}']"
                break
        print(f"  {c['id']}  {c['title']!r}{marker}")


# ============================================================
if __name__ == "__main__":
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("mode", choices=["populate", "sync", "migrate-task-column",
                                    "stamp-locations", "populate-slugs", "apply-verified",
                                    "apply-template", "apply-hyperlinks",
                                    "list-rollout-cols"])
    p.add_argument("--execute", action="store_true",
                   help="actually write changes (default is dry run)")
    p.add_argument("--cell-links", action="store_true",
                   help="convert direct-source cells to live cell links")
    p.add_argument("--excel", default=None,
                   help="override the Excel path for populate-slugs")
    args = p.parse_args()
    if args.mode == "populate":
        mode_populate(args.execute, args.cell_links)
    elif args.mode == "sync":
        mode_sync(args.execute, args.cell_links)
    elif args.mode == "migrate-task-column":
        mode_migrate(args.execute)
    elif args.mode == "populate-slugs":
        mode_populate_slugs(args.execute, args.excel)
    elif args.mode == "apply-verified":
        mode_apply_verified(args.execute)
    elif args.mode == "apply-template":
        mode_apply_template(args.execute)
    elif args.mode == "apply-hyperlinks":
        mode_apply_hyperlinks(args.execute)
    elif args.mode == "list-rollout-cols":
        mode_list_rollout_cols()
    else:
        mode_stamp_locations(args.execute)
